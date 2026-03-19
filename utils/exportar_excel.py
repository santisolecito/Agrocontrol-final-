import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_VERDE1 = "2C5F2D"
_VERDE2 = "4A7C43"
_VERDE3 = "EEF4EE"
_BORDE  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

def _ventana_filtros(titulo, columnas, filas):
    resultado = {"filas": None, "desc": None}
    win = tk.Toplevel()
    win.title(f"Filtros — Exportar {titulo}")
    win.geometry("480x360")
    win.resizable(False, False)
    win.configure(bg="#f0f4f0")
    win.grab_set()
    BG = "#f0f4f0"

    header = tk.Frame(win, bg="#2d5a27", height=45)
    header.pack(fill="x")
    header.pack_propagate(False)
    tk.Label(header, text=f"📊 Exportar {titulo} a Excel",
             font=("Segoe UI", 12, "bold"), bg="#2d5a27", fg="white"
             ).pack(pady=10)

    frm = tk.Frame(win, bg=BG)
    frm.pack(fill="x", padx=20, pady=10)

    tk.Label(frm, text="Filtrar por columna:", font=("Segoe UI", 10, "bold"), bg=BG
             ).grid(row=0, column=0, sticky="w", pady=4)
    col_var = tk.StringVar(value="(Sin filtro)")
    combo_col = ttk.Combobox(frm, textvariable=col_var,
                              values=["(Sin filtro)"] + list(columnas),
                              state="readonly", width=22)
    combo_col.grid(row=0, column=1, sticky="w", padx=8)

    tk.Label(frm, text="Valor a buscar:", font=("Segoe UI", 10, "bold"), bg=BG
             ).grid(row=1, column=0, sticky="w", pady=4)
    val_entry = tk.Entry(frm, width=25, font=("Segoe UI", 10), relief="solid", bd=1)
    val_entry.grid(row=1, column=1, sticky="w", padx=8)

    tk.Label(frm, text="─"*50, bg=BG, fg="#cccccc"
             ).grid(row=2, column=0, columnspan=2, pady=4)

    tk.Label(frm, text="Columna de fecha:", font=("Segoe UI", 10, "bold"), bg=BG
             ).grid(row=3, column=0, sticky="w", pady=4)
    fecha_col_var = tk.StringVar(value="(Ninguna)")
    combo_fecha = ttk.Combobox(frm, textvariable=fecha_col_var,
                                values=["(Ninguna)"] + list(columnas),
                                state="readonly", width=22)
    combo_fecha.grid(row=3, column=1, sticky="w", padx=8)

    tk.Label(frm, text="Desde (YYYY-MM-DD):", font=("Segoe UI", 10, "bold"), bg=BG
             ).grid(row=4, column=0, sticky="w", pady=4)
    desde_entry = tk.Entry(frm, width=16, font=("Segoe UI", 10), relief="solid", bd=1)
    desde_entry.grid(row=4, column=1, sticky="w", padx=8)

    tk.Label(frm, text="Hasta (YYYY-MM-DD):", font=("Segoe UI", 10, "bold"), bg=BG
             ).grid(row=5, column=0, sticky="w", pady=4)
    hasta_entry = tk.Entry(frm, width=16, font=("Segoe UI", 10), relief="solid", bd=1)
    hasta_entry.grid(row=5, column=1, sticky="w", padx=8)

    def aplicar():
        filas_f = list(filas)
        desc_parts = []
        col_sel = col_var.get()
        val_sel = val_entry.get().strip().lower()
        if col_sel != "(Sin filtro)" and val_sel:
            idx = list(columnas).index(col_sel)
            filas_f = [f for f in filas_f if val_sel in str(f[idx]).lower()]
            desc_parts.append(f"{col_sel}='{val_entry.get().strip()}'")
        col_fecha = fecha_col_var.get()
        desde_str = desde_entry.get().strip()
        hasta_str = hasta_entry.get().strip()
        if col_fecha != "(Ninguna)" and (desde_str or hasta_str):
            idx_f = list(columnas).index(col_fecha)
            def en_rango(fila):
                try:
                    val = str(fila[idx_f])[:10]
                    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
                        try:
                            fecha = datetime.strptime(val, fmt).date(); break
                        except ValueError: continue
                    else: return True
                    if desde_str:
                        if fecha < datetime.strptime(desde_str, "%Y-%m-%d").date(): return False
                    if hasta_str:
                        if fecha > datetime.strptime(hasta_str, "%Y-%m-%d").date(): return False
                    return True
                except Exception: return True
            filas_f = [f for f in filas_f if en_rango(f)]
            if desde_str: desc_parts.append(f"desde {desde_str}")
            if hasta_str: desc_parts.append(f"hasta {hasta_str}")
        if not filas_f:
            messagebox.showwarning("Sin resultados", "El filtro no devuelve registros."); return
        resultado["filas"] = filas_f
        resultado["desc"]  = "  |  Filtro: " + ", ".join(desc_parts) if desc_parts else ""
        win.destroy()

    def sin_filtro():
        resultado["filas"] = list(filas)
        resultado["desc"]  = ""
        win.destroy()

    btn_frame = tk.Frame(win, bg=BG)
    btn_frame.pack(side="bottom", pady=14)
    for txt, bg, cmd in [
        ("📤 Exportar con filtro", "#4CAF50", aplicar),
        ("📋 Exportar todo",       "#2196F3", sin_filtro),
        ("✖ Cancelar",             "#9E9E9E", win.destroy),
    ]:
        tk.Button(btn_frame, text=txt, font=("Segoe UI", 10, "bold"),
                  bg=bg, fg="white", padx=10, pady=5, relief="flat",
                  cursor="hand2", command=cmd).pack(side="left", padx=5)

    win.wait_window()
    return resultado["filas"], resultado["desc"]


def exportar_excel(titulo, columnas, filas):
    filas_f, desc_filtro = _ventana_filtros(titulo, columnas, filas)
    if filas_f is None:
        return
    ruta = filedialog.asksaveasfilename(
        defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
        title=f"Guardar {titulo}.xlsx",
        initialfile=f"{titulo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    if not ruta:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo[:31]
    ncols = len(columnas)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=f"AgroControl — {titulo}{desc_filtro}")
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=_VERDE1)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c2 = ws.cell(row=2, column=1,
                 value=f"Exportado: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   Registros: {len(filas_f)}")
    c2.font = Font(italic=True, size=9, color="555555")
    c2.alignment = Alignment(horizontal="center")
    for ci, nombre in enumerate(columnas, start=1):
        c = ws.cell(row=3, column=ci, value=nombre)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=_VERDE2)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDE
    ws.row_dimensions[3].height = 20
    for ri, fila in enumerate(filas_f, start=4):
        bg = "FFFFFF" if ri % 2 == 0 else _VERDE3
        for ci, valor in enumerate(fila, start=1):
            c = ws.cell(row=ri, column=ci, value=str(valor) if valor is not None else "")
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(vertical="center")
            c.border = _BORDE
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 42)
    ws.freeze_panes = "A4"
    wb.save(ruta)
    messagebox.showinfo("Excel exportado", f"✅ {len(filas_f)} registros guardados en:\n{ruta}")
